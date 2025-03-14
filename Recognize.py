import pandas as pd
import numpy as np
import cv2 as cv
from openpyxl import Workbook, load_workbook
import xlrd as xl
from datetime import date

wb = load_workbook('Attendance.xlsx')
ws = wb.active

# find a column 
col = ws.max_column

# put a date on the top of it
date = date.today()
date = date.strftime("%d-%b-%y")
date_col = chr(65+col)
ws[date_col+str(1)] = date

id_names = pd.read_csv('id-names.csv')
id_names = id_names[['id', 'name']]
faceClassifier = cv.CascadeClassifier('Classifiers/haarface.xml')

lbph = cv.face.LBPHFaceRecognizer_create()
lbph.read('Classifiers/TrainedLBPH.yml')

camera = cv.VideoCapture(0)

while cv.waitKey(1) & 0xFF != ord('q'):
    _, img = camera.read()
    grey = cv.cvtColor(img, cv.COLOR_BGR2GRAY)

    faces = faceClassifier.detectMultiScale(grey, scaleFactor=1.2, minNeighbors=4)

    for x, y, w, h in faces:
        faceRegion = grey[y:y + h, x:x + w]
        faceRegion = cv.resize(faceRegion, (220, 220))

        label, trust = lbph.predict(faceRegion)
        try:
            name = id_names[id_names['id'] == label]['name'].item()
            cv.rectangle(img, (x, y), (x + w, y + h), (0, 255, 0), 2)
            cv.putText(img, name, (x, y + h + 30), cv.FONT_HERSHEY_COMPLEX, 1, (0, 255, 0))
            # mark present in front of the name using ID
            ws[date_col+str(label+1)] = "P"
        except:
            pass

    cv.imshow('Recognize', img)

camera.release()
cv.destroyAllWindows()

# mark absent to those who are not marked
row = ws.max_row

r=2
while(r<row):
    c=date_col + str(r)
    if( str(ws[c].value) == "None"):
        ws[c] = "A"
    r = r+1


# Save the excel sheet
wb.save('Attendance.xlsx')
